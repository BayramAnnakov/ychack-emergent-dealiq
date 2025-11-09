#!/usr/bin/env python3
"""
Excel Output Validator Agent using Claude Agent SDK

This agent validates Excel files and uses Claude to provide intelligent analysis
of issues and suggestions for fixes.
"""
import os
import json
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field, asdict
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.agents.base import BaseAgent


logger = logging.getLogger(__name__)


@dataclass
class ValidationIssue:
    """Represents a validation issue found in an Excel file"""
    severity: str  # 'critical', 'warning', 'info'
    category: str  # 'formula_error', 'data_quality', 'formatting', etc.
    sheet: str
    cell: str
    message: str
    value: Any = None
    formula: Optional[str] = None


@dataclass
class ValidationReport:
    """Complete validation report for an Excel file"""
    file_path: str
    is_valid: bool
    total_issues: int
    critical_issues: int
    warnings: int
    info_messages: int
    issues: List[ValidationIssue] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)
    ai_analysis: Optional[str] = None

    def add_issue(self, issue: ValidationIssue):
        """Add an issue to the report"""
        self.issues.append(issue)
        self.total_issues += 1

        if issue.severity == 'critical':
            self.critical_issues += 1
            self.is_valid = False
        elif issue.severity == 'warning':
            self.warnings += 1
        elif issue.severity == 'info':
            self.info_messages += 1

    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary"""
        return {
            'file_path': self.file_path,
            'is_valid': self.is_valid,
            'total_issues': self.total_issues,
            'critical_issues': self.critical_issues,
            'warnings': self.warnings,
            'info_messages': self.info_messages,
            'issues': [asdict(issue) for issue in self.issues],
            'summary': self.summary,
            'ai_analysis': self.ai_analysis
        }

    def print_report(self, verbose: bool = True):
        """Print a human-readable report"""
        print(f"\n{'='*80}")
        print(f"📋 EXCEL VALIDATION REPORT")
        print(f"{'='*80}")
        print(f"File: {self.file_path}")
        print(f"Status: {'✅ VALID' if self.is_valid else '❌ INVALID'}")
        print(f"\n📊 Issue Summary:")
        print(f"   Critical: {self.critical_issues}")
        print(f"   Warnings: {self.warnings}")
        print(f"   Info: {self.info_messages}")
        print(f"   Total: {self.total_issues}")

        if self.summary:
            print(f"\n📈 Statistics:")
            for key, value in self.summary.items():
                print(f"   {key}: {value}")

        if self.ai_analysis:
            print(f"\n🤖 AI Analysis:")
            print(f"{self.ai_analysis}")

        if self.issues and verbose:
            print(f"\n🔍 Issues Found:")

            # Group by severity
            for severity in ['critical', 'warning', 'info']:
                severity_issues = [i for i in self.issues if i.severity == severity]
                if severity_issues:
                    emoji = '🔴' if severity == 'critical' else '🟡' if severity == 'warning' else 'ℹ️'
                    print(f"\n{emoji} {severity.upper()} ({len(severity_issues)}):")
                    for issue in severity_issues:
                        print(f"   [{issue.sheet}!{issue.cell}] {issue.message}")
                        if issue.formula:
                            print(f"      Formula: {issue.formula}")
                        if issue.value is not None:
                            print(f"      Value: {issue.value}")

        print(f"\n{'='*80}")


class ExcelValidatorAgent(BaseAgent):
    """
    Excel Validator Agent that uses Claude to intelligently analyze Excel files
    """

    # Known Excel error values
    ERROR_VALUES = {
        '#DIV/0!': 'Division by zero',
        '#VALUE!': 'Wrong type of argument or operand',
        '#REF!': 'Invalid cell reference',
        '#N/A': 'Value not available',
        '#NAME?': 'Unrecognized formula name',
        '#NULL!': 'Invalid intersection of ranges',
        '#NUM!': 'Invalid numeric value'
    }

    def __init__(self):
        super().__init__(
            name="Excel Validator",
            description="analyzing and validating Excel spreadsheet outputs for quality, correctness, and identifying potential issues"
        )

    async def process(self, data: Any, context: Optional[Dict] = None) -> Dict[str, Any]:
        """
        Process and validate an Excel file

        Args:
            data: Either file path (str) or dict with 'file_path' key
            context: Optional context with validation parameters

        Returns:
            Dictionary containing validation report
        """
        # Extract file path
        if isinstance(data, str):
            file_path = data
        elif isinstance(data, dict) and 'file_path' in data:
            file_path = data['file_path']
        else:
            return {
                'error': 'Invalid input: expected file path or dict with file_path',
                'is_valid': False
            }

        # Run validation
        report = self._scan_file(file_path)

        # If there are issues, use Claude to analyze them
        if report.total_issues > 0:
            ai_analysis = await self._get_ai_analysis(report)
            report.ai_analysis = ai_analysis

        return report.to_dict()

    def _scan_file(self, file_path: str) -> ValidationReport:
        """
        Scan Excel file for issues (non-AI validation)

        Args:
            file_path: Path to Excel file

        Returns:
            ValidationReport with all issues found
        """
        report = ValidationReport(
            file_path=file_path,
            is_valid=True,
            total_issues=0,
            critical_issues=0,
            warnings=0,
            info_messages=0
        )

        # Check if file exists
        if not os.path.exists(file_path):
            report.add_issue(ValidationIssue(
                severity='critical',
                category='file',
                sheet='N/A',
                cell='N/A',
                message=f"File not found: {file_path}"
            ))
            return report

        try:
            # Load workbook twice - once for formulas, once for calculated values
            wb = load_workbook(file_path, data_only=False)
            wb_data = load_workbook(file_path, data_only=True)

            # Track statistics
            total_cells = 0
            formula_cells = 0
            error_cells = 0
            empty_cells = 0

            # Validate each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = wb_data[sheet_name]

                logger.info(f"Validating sheet: {sheet_name}")

                # Check each cell
                for row_idx, row in enumerate(sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        total_cells += 1

                        # Get corresponding data cell for calculated value
                        data_cell = sheet_data.cell(row=row_idx, column=col_idx)

                        # Check for formula errors in calculated values
                        if isinstance(data_cell.value, str) and data_cell.value in self.ERROR_VALUES:
                            error_cells += 1
                            formula = str(cell.value) if hasattr(cell, 'value') and str(cell.value).startswith('=') else None
                            report.add_issue(ValidationIssue(
                                severity='critical',
                                category='formula_error',
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                message=f"Formula error: {self.ERROR_VALUES[data_cell.value]}",
                                value=data_cell.value,
                                formula=formula
                            ))

                        # Check if cell has a formula
                        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells += 1

                            # Check for potential division by zero in formula
                            formula = str(cell.value)
                            if '/' in formula and self._might_divide_by_zero(formula):
                                report.add_issue(ValidationIssue(
                                    severity='warning',
                                    category='formula_quality',
                                    sheet=sheet_name,
                                    cell=cell.coordinate,
                                    message="Formula may result in division by zero",
                                    formula=formula
                                ))

                        # Check for empty cells
                        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                            empty_cells += 1

            # Update summary statistics
            report.summary = {
                'total_sheets': len(wb.sheetnames),
                'total_cells': total_cells,
                'formula_cells': formula_cells,
                'error_cells': error_cells,
                'empty_cells': empty_cells,
                'sheets': wb.sheetnames
            }

            # Add info messages
            if formula_cells > 0:
                report.add_issue(ValidationIssue(
                    severity='info',
                    category='statistics',
                    sheet='All',
                    cell='N/A',
                    message=f"Found {formula_cells} cells with formulas"
                ))

            if error_cells == 0:
                report.add_issue(ValidationIssue(
                    severity='info',
                    category='quality',
                    sheet='All',
                    cell='N/A',
                    message="No formula errors detected"
                ))

            wb_data.close()
            wb.close()

        except InvalidFileException as e:
            report.add_issue(ValidationIssue(
                severity='critical',
                category='file',
                sheet='N/A',
                cell='N/A',
                message=f"Invalid Excel file: {str(e)}"
            ))
        except Exception as e:
            report.add_issue(ValidationIssue(
                severity='critical',
                category='error',
                sheet='N/A',
                cell='N/A',
                message=f"Error validating file: {str(e)}"
            ))

        return report

    def _might_divide_by_zero(self, formula: str) -> bool:
        """
        Check if a formula might result in division by zero
        This is a simple heuristic check, not exhaustive
        """
        risky_patterns = [
            '/0',  # Direct division by zero
            '/(0)',  # Division by zero in parentheses
        ]

        for pattern in risky_patterns:
            if pattern in formula:
                return True

        return False

    async def _get_ai_analysis(self, report: ValidationReport) -> str:
        """
        Use Claude to analyze validation issues and provide recommendations

        Args:
            report: ValidationReport with issues

        Returns:
            AI-generated analysis and recommendations
        """
        # Build analysis prompt
        issues_summary = {
            'file': report.file_path,
            'statistics': report.summary,
            'critical_issues': report.critical_issues,
            'warnings': report.warnings,
            'issues': []
        }

        # Add issues details (limit to avoid token overflow)
        for issue in report.issues[:20]:  # Limit to first 20 issues
            issues_summary['issues'].append({
                'severity': issue.severity,
                'category': issue.category,
                'sheet': issue.sheet,
                'cell': issue.cell,
                'message': issue.message,
                'formula': issue.formula,
                'value': str(issue.value) if issue.value else None
            })

        prompt = f"""Analyze this Excel file validation report and provide insights:

{json.dumps(issues_summary, indent=2)}

Please provide:
1. A brief summary of the validation results
2. Analysis of critical issues (if any) and their potential causes
3. Recommendations for fixing the issues
4. Assessment of overall data quality

Keep your response concise and actionable."""

        try:
            analysis = await self.query_single(prompt)
            return analysis
        except Exception as e:
            logger.error(f"Error getting AI analysis: {e}")
            return f"AI analysis unavailable: {str(e)}"


async def validate_excel_with_ai(file_path: str, verbose: bool = True) -> ValidationReport:
    """
    Convenience function to validate an Excel file with AI analysis

    Args:
        file_path: Path to Excel file
        verbose: Print detailed report

    Returns:
        ValidationReport with AI analysis
    """
    async with ExcelValidatorAgent() as agent:
        result = await agent.process(file_path)

        # Convert back to ValidationReport for printing
        report = ValidationReport(
            file_path=result['file_path'],
            is_valid=result['is_valid'],
            total_issues=result['total_issues'],
            critical_issues=result['critical_issues'],
            warnings=result['warnings'],
            info_messages=result['info_messages'],
            issues=[ValidationIssue(**issue) for issue in result['issues']],
            summary=result['summary'],
            ai_analysis=result.get('ai_analysis')
        )

        if verbose:
            report.print_report(verbose=True)

        return report


if __name__ == "__main__":
    """Test the validator agent"""
    import sys
    import asyncio

    if len(sys.argv) < 2:
        print("Usage: python excel_validator_agent.py <excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]

    async def main():
        report = await validate_excel_with_ai(file_path, verbose=True)
        return 0 if report.is_valid else 1

    exit_code = asyncio.run(main())
    sys.exit(exit_code)
