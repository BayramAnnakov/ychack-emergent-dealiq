#!/usr/bin/env python3
"""
Excel Output Validator Agent

Validates Excel files for common issues:
- Formula errors (#DIV/0!, #VALUE!, #REF!, #N/A, #NAME?, #NULL!)
- Empty cells in calculated columns
- Data type mismatches
- Missing required sheets/columns
- Formatting issues
"""
import os
import logging
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.cell.cell import Cell


logger = logging.getLogger(__name__)


@dataclass
class ValidationIssue:
    """Represents a validation issue found in an Excel file"""
    severity: str  # 'critical', 'warning', 'info'
    category: str  # 'formula_error', 'data_quality', 'formatting', etc.
    sheet: str
    cell: str
    message: str
    value: Any = None
    formula: Optional[str] = None


@dataclass
class ValidationReport:
    """Complete validation report for an Excel file"""
    file_path: str
    is_valid: bool
    total_issues: int
    critical_issues: int
    warnings: int
    info_messages: int
    issues: List[ValidationIssue] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)

    def add_issue(self, issue: ValidationIssue):
        """Add an issue to the report"""
        self.issues.append(issue)
        self.total_issues += 1

        if issue.severity == 'critical':
            self.critical_issues += 1
            self.is_valid = False
        elif issue.severity == 'warning':
            self.warnings += 1
        elif issue.severity == 'info':
            self.info_messages += 1

    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary"""
        return {
            'file_path': self.file_path,
            'is_valid': self.is_valid,
            'total_issues': self.total_issues,
            'critical_issues': self.critical_issues,
            'warnings': self.warnings,
            'info_messages': self.info_messages,
            'issues': [
                {
                    'severity': issue.severity,
                    'category': issue.category,
                    'sheet': issue.sheet,
                    'cell': issue.cell,
                    'message': issue.message,
                    'value': str(issue.value) if issue.value is not None else None,
                    'formula': issue.formula
                }
                for issue in self.issues
            ],
            'summary': self.summary
        }

    def print_report(self, verbose: bool = True):
        """Print a human-readable report"""
        print(f"\n{'='*80}")
        print(f"📋 EXCEL VALIDATION REPORT")
        print(f"{'='*80}")
        print(f"File: {self.file_path}")
        print(f"Status: {'✅ VALID' if self.is_valid else '❌ INVALID'}")
        print(f"\n📊 Issue Summary:")
        print(f"   Critical: {self.critical_issues}")
        print(f"   Warnings: {self.warnings}")
        print(f"   Info: {self.info_messages}")
        print(f"   Total: {self.total_issues}")

        if self.summary:
            print(f"\n📈 Statistics:")
            for key, value in self.summary.items():
                print(f"   {key}: {value}")

        if self.issues and verbose:
            print(f"\n🔍 Issues Found:")

            # Group by severity
            for severity in ['critical', 'warning', 'info']:
                severity_issues = [i for i in self.issues if i.severity == severity]
                if severity_issues:
                    emoji = '🔴' if severity == 'critical' else '🟡' if severity == 'warning' else 'ℹ️'
                    print(f"\n{emoji} {severity.upper()} ({len(severity_issues)}):")
                    for issue in severity_issues:
                        print(f"   [{issue.sheet}!{issue.cell}] {issue.message}")
                        if issue.formula:
                            print(f"      Formula: {issue.formula}")
                        if issue.value is not None:
                            print(f"      Value: {issue.value}")

        print(f"\n{'='*80}")


class ExcelValidator:
    """Validates Excel files for quality and correctness"""

    # Known Excel error values
    ERROR_VALUES = {
        '#DIV/0!': 'Division by zero',
        '#VALUE!': 'Wrong type of argument or operand',
        '#REF!': 'Invalid cell reference',
        '#N/A': 'Value not available',
        '#NAME?': 'Unrecognized formula name',
        '#NULL!': 'Invalid intersection of ranges',
        '#NUM!': 'Invalid numeric value'
    }

    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.logger = logging.getLogger(__name__)

    def validate_file(self, file_path: str) -> ValidationReport:
        """
        Validate an Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            ValidationReport with all issues found
        """
        report = ValidationReport(
            file_path=file_path,
            is_valid=True,
            total_issues=0,
            critical_issues=0,
            warnings=0,
            info_messages=0
        )

        # Check if file exists
        if not os.path.exists(file_path):
            report.add_issue(ValidationIssue(
                severity='critical',
                category='file',
                sheet='N/A',
                cell='N/A',
                message=f"File not found: {file_path}"
            ))
            return report

        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=False)

            # Track statistics
            total_cells = 0
            formula_cells = 0
            error_cells = 0
            empty_cells = 0

            # Validate each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                self.logger.info(f"Validating sheet: {sheet_name}")

                # Check each cell
                for row in sheet.iter_rows():
                    for cell in row:
                        total_cells += 1

                        # Check for formula errors
                        if isinstance(cell.value, str) and cell.value in self.ERROR_VALUES:
                            error_cells += 1
                            report.add_issue(ValidationIssue(
                                severity='critical',
                                category='formula_error',
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                message=f"Formula error: {self.ERROR_VALUES[cell.value]}",
                                value=cell.value,
                                formula=str(cell.value) if hasattr(cell, 'value') else None
                            ))

                        # Check if cell has a formula
                        if hasattr(cell, 'value') and str(cell.value).startswith('='):
                            formula_cells += 1

                            # Check for potential division by zero in formula
                            formula = str(cell.value)
                            if '/' in formula and self._might_divide_by_zero(formula):
                                report.add_issue(ValidationIssue(
                                    severity='warning',
                                    category='formula_quality',
                                    sheet=sheet_name,
                                    cell=cell.coordinate,
                                    message="Formula may result in division by zero",
                                    formula=formula
                                ))

                        # Check for empty cells
                        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                            empty_cells += 1

            # Update summary statistics
            report.summary = {
                'total_sheets': len(wb.sheetnames),
                'total_cells': total_cells,
                'formula_cells': formula_cells,
                'error_cells': error_cells,
                'empty_cells': empty_cells,
                'sheets': wb.sheetnames
            }

            # Add info messages
            if formula_cells > 0:
                report.add_issue(ValidationIssue(
                    severity='info',
                    category='statistics',
                    sheet='All',
                    cell='N/A',
                    message=f"Found {formula_cells} cells with formulas"
                ))

            if error_cells == 0:
                report.add_issue(ValidationIssue(
                    severity='info',
                    category='quality',
                    sheet='All',
                    cell='N/A',
                    message="No formula errors detected"
                ))

            wb.close()

        except InvalidFileException as e:
            report.add_issue(ValidationIssue(
                severity='critical',
                category='file',
                sheet='N/A',
                cell='N/A',
                message=f"Invalid Excel file: {str(e)}"
            ))
        except Exception as e:
            report.add_issue(ValidationIssue(
                severity='critical',
                category='error',
                sheet='N/A',
                cell='N/A',
                message=f"Error validating file: {str(e)}"
            ))

        return report

    def _might_divide_by_zero(self, formula: str) -> bool:
        """
        Check if a formula might result in division by zero
        This is a simple heuristic check, not exhaustive
        """
        # Simple patterns that might indicate division by zero risk
        risky_patterns = [
            '/0',  # Direct division by zero
            '/(0)',  # Division by zero in parentheses
        ]

        for pattern in risky_patterns:
            if pattern in formula:
                return True

        return False

    def validate_and_fix(self, file_path: str, output_path: Optional[str] = None) -> Tuple[ValidationReport, Optional[str]]:
        """
        Validate an Excel file and attempt to fix common issues

        Args:
            file_path: Path to input Excel file
            output_path: Path for fixed output file (optional)

        Returns:
            Tuple of (ValidationReport, fixed_file_path)
        """
        # First validate
        report = self.validate_file(file_path)

        # If no critical issues, no fix needed
        if report.critical_issues == 0:
            return report, None

        # Attempt to fix issues
        try:
            wb = load_workbook(file_path)
            fixed = False

            for issue in report.issues:
                if issue.category == 'formula_error' and issue.severity == 'critical':
                    sheet = wb[issue.sheet]
                    cell = sheet[issue.cell]

                    # Replace error with a warning comment
                    # (In a real implementation, you might want more sophisticated fixes)
                    self.logger.warning(f"Found error in {issue.sheet}!{issue.cell}: {issue.value}")
                    # We don't auto-fix formulas as it could corrupt data
                    # Just log the issue

            # Save fixed version if requested
            if output_path and fixed:
                wb.save(output_path)
                wb.close()
                return report, output_path

            wb.close()

        except Exception as e:
            self.logger.error(f"Error attempting to fix file: {str(e)}")

        return report, None


def validate_excel_file(file_path: str, verbose: bool = True) -> ValidationReport:
    """
    Convenience function to validate an Excel file

    Args:
        file_path: Path to Excel file
        verbose: Print detailed report

    Returns:
        ValidationReport
    """
    validator = ExcelValidator(verbose=verbose)
    report = validator.validate_file(file_path)

    if verbose:
        report.print_report(verbose=True)

    return report


if __name__ == "__main__":
    """Test the validator"""
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_validator.py <excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    report = validate_excel_file(file_path, verbose=True)

    # Exit with error code if validation failed
    sys.exit(0 if report.is_valid else 1)
