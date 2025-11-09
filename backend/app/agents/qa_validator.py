#!/usr/bin/env python3
"""
Comprehensive QA Validator for Excel and PDF outputs
Validates quality, detects errors, and provides actionable feedback
"""
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import PyPDF2

logger = logging.getLogger(__name__)


@dataclass
class QAIssue:
    """Represents a quality issue found in output file"""
    severity: str  # 'critical', 'warning', 'info'
    category: str  # 'formula_error', 'formatting', 'data_quality', etc.
    location: str  # Sheet!Cell or Page number
    message: str
    value: Any = None
    formula: Optional[str] = None
    suggestion: Optional[str] = None


@dataclass
class QAReport:
    """Complete QA validation report"""
    file_path: str
    file_type: str  # 'excel' or 'pdf'
    is_valid: bool
    quality_score: int  # 0-100
    total_issues: int
    critical_issues: int
    warnings: int
    info_messages: int
    issues: List[QAIssue] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)

    def add_issue(self, issue: QAIssue):
        """Add an issue and update counts"""
        self.issues.append(issue)
        self.total_issues += 1

        if issue.severity == 'critical':
            self.critical_issues += 1
            self.is_valid = False
        elif issue.severity == 'warning':
            self.warnings += 1
        elif issue.severity == 'info':
            self.info_messages += 1

    def calculate_quality_score(self):
        """Calculate quality score 0-100 based on issues"""
        if self.critical_issues > 0:
            base_score = max(0, 50 - (self.critical_issues * 10))
        else:
            base_score = 100 - (self.warnings * 5) - (self.info_messages * 1)
        
        self.quality_score = max(0, min(100, base_score))

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary"""
        return {
            'file_path': self.file_path,
            'file_type': self.file_type,
            'is_valid': self.is_valid,
            'quality_score': self.quality_score,
            'total_issues': self.total_issues,
            'critical_issues': self.critical_issues,
            'warnings': self.warnings,
            'info_messages': self.info_messages,
            'issues': [
                {
                    'severity': issue.severity,
                    'category': issue.category,
                    'location': issue.location,
                    'message': issue.message,
                    'value': str(issue.value) if issue.value is not None else None,
                    'formula': issue.formula,
                    'suggestion': issue.suggestion
                }
                for issue in self.issues
            ],
            'summary': self.summary
        }


class QAValidator:
    """Quality Assurance validator for Excel and PDF files"""

    # Excel error values
    EXCEL_ERRORS = {
        '#DIV/0!': 'Division by zero',
        '#VALUE!': 'Wrong type of argument or operand',
        '#REF!': 'Invalid cell reference',
        '#N/A': 'Value not available',
        '#NAME?': 'Unrecognized formula name',
        '#NULL!': 'Invalid intersection of ranges',
        '#NUM!': 'Invalid numeric value'
    }

    def __init__(self, verbose: bool = True):
        self.verbose = verbose

    def validate_file(self, file_path: str) -> QAReport:
        """Main validation entry point - detects file type and validates"""
        if not os.path.exists(file_path):
            report = QAReport(
                file_path=file_path,
                file_type='unknown',
                is_valid=False,
                quality_score=0,
                total_issues=1,
                critical_issues=1,
                warnings=0,
                info_messages=0
            )
            report.add_issue(QAIssue(
                severity='critical',
                category='file',
                location='N/A',
                message='File not found'
            ))
            return report

        # Detect file type
        if file_path.endswith('.xlsx') or file_path.endswith('.xlsm'):
            return self._validate_excel(file_path)
        elif file_path.endswith('.pdf'):
            return self._validate_pdf(file_path)
        else:
            report = QAReport(
                file_path=file_path,
                file_type='unknown',
                is_valid=False,
                quality_score=0,
                total_issues=1,
                critical_issues=1,
                warnings=0,
                info_messages=0
            )
            report.add_issue(QAIssue(
                severity='critical',
                category='file',
                location='N/A',
                message='Unsupported file type'
            ))
            return report

    def _validate_excel(self, file_path: str) -> QAReport:
        """Validate Excel file comprehensively"""
        report = QAReport(
            file_path=file_path,
            file_type='excel',
            is_valid=True,
            quality_score=100,
            total_issues=0,
            critical_issues=0,
            warnings=0,
            info_messages=0
        )

        try:
            # Load workbook - both with formulas and calculated values
            wb = load_workbook(file_path, data_only=False)
            wb_data = load_workbook(file_path, data_only=True)

            total_cells = 0
            formula_cells = 0
            error_cells = 0
            empty_cells = 0
            numeric_cells = 0

            # Validate each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = wb_data[sheet_name]

                for row_idx, row in enumerate(sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        total_cells += 1
                        data_cell = sheet_data.cell(row=row_idx, column=col_idx)

                        # Check for formula errors
                        if isinstance(data_cell.value, str) and data_cell.value in self.EXCEL_ERRORS:
                            error_cells += 1
                            formula = str(cell.value) if str(cell.value).startswith('=') else None
                            
                            report.add_issue(QAIssue(
                                severity='critical',
                                category='formula_error',
                                location=f"{sheet_name}!{cell.coordinate}",
                                message=f"{self.EXCEL_ERRORS[data_cell.value]}",
                                value=data_cell.value,
                                formula=formula,
                                suggestion=self._suggest_formula_fix(data_cell.value, formula)
                            ))

                        # Check if has formula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells += 1

                            # Check for division by zero risk
                            if self._has_div_zero_risk(cell.value):
                                report.add_issue(QAIssue(
                                    severity='warning',
                                    category='formula_quality',
                                    location=f"{sheet_name}!{cell.coordinate}",
                                    message="Potential division by zero risk",
                                    formula=cell.value,
                                    suggestion="Consider using IFERROR() or IF() to handle zero divisors"
                                ))

                        # Track numeric cells
                        if isinstance(data_cell.value, (int, float)):
                            numeric_cells += 1

                        # Check for empty cells
                        if cell.value is None:
                            empty_cells += 1

            # Statistics
            report.summary = {
                'total_sheets': len(wb.sheetnames),
                'total_cells': total_cells,
                'formula_cells': formula_cells,
                'error_cells': error_cells,
                'empty_cells': empty_cells,
                'numeric_cells': numeric_cells,
                'formula_percentage': round((formula_cells / total_cells * 100) if total_cells > 0 else 0, 1)
            }

            # Quality indicators
            if formula_cells > 0:
                report.add_issue(QAIssue(
                    severity='info',
                    category='quality',
                    location='All Sheets',
                    message=f"✅ {formula_cells} formula-driven calculations found",
                    suggestion="Formula-based analysis is best practice"
                ))

            if error_cells == 0:
                report.add_issue(QAIssue(
                    severity='info',
                    category='quality',
                    location='All Sheets',
                    message="✅ Zero formula errors - excellent quality"
                ))

            wb_data.close()
            wb.close()

        except Exception as e:
            report.add_issue(QAIssue(
                severity='critical',
                category='error',
                location='File',
                message=f"Validation error: {str(e)}"
            ))

        report.calculate_quality_score()
        return report

    def _validate_pdf(self, file_path: str) -> QAReport:
        """Validate PDF file"""
        report = QAReport(
            file_path=file_path,
            file_type='pdf',
            is_valid=True,
            quality_score=100,
            total_issues=0,
            critical_issues=0,
            warnings=0,
            info_messages=0
        )

        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                page_count = len(pdf_reader.pages)
                file_size = os.path.getsize(file_path)
                
                # Basic checks
                if page_count == 0:
                    report.add_issue(QAIssue(
                        severity='critical',
                        category='content',
                        location='Document',
                        message='PDF has no pages',
                        suggestion='Regenerate the PDF document'
                    ))
                
                # Extract text to verify content
                total_text = ""
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    try:
                        text = page.extract_text()
                        total_text += text
                        
                        # Check for empty pages
                        if not text or len(text.strip()) < 10:
                            report.add_issue(QAIssue(
                                severity='warning',
                                category='content',
                                location=f'Page {page_num}',
                                message='Page appears to be empty or has minimal content',
                                suggestion='Verify page content is rendering correctly'
                            ))
                    except Exception as e:
                        report.add_issue(QAIssue(
                            severity='warning',
                            category='extraction',
                            location=f'Page {page_num}',
                            message=f'Could not extract text: {str(e)}'
                        ))
                
                # Statistics
                report.summary = {
                    'total_pages': page_count,
                    'file_size_kb': round(file_size / 1024, 1),
                    'total_characters': len(total_text),
                    'has_metadata': pdf_reader.metadata is not None
                }
                
                # Quality checks
                if page_count > 0:
                    report.add_issue(QAIssue(
                        severity='info',
                        category='quality',
                        location='Document',
                        message=f"✅ {page_count} page{'s' if page_count > 1 else ''} generated successfully"
                    ))
                
                if len(total_text) > 100:
                    report.add_issue(QAIssue(
                        severity='info',
                        category='quality',
                        location='Document',
                        message=f"✅ {len(total_text)} characters of content extracted"
                    ))
                else:
                    report.add_issue(QAIssue(
                        severity='warning',
                        category='content',
                        location='Document',
                        message='PDF has very little text content',
                        suggestion='Verify the document generated correctly'
                    ))

        except Exception as e:
            report.add_issue(QAIssue(
                severity='critical',
                category='error',
                location='File',
                message=f"PDF validation error: {str(e)}"
            ))

        report.calculate_quality_score()
        return report

    def _has_div_zero_risk(self, formula: str) -> bool:
        """Check if formula has division by zero risk"""
        risky_patterns = ['/0', '/(0)', '/ 0']
        return any(pattern in formula for pattern in risky_patterns)

    def _suggest_formula_fix(self, error_value: str, formula: Optional[str]) -> str:
        """Suggest fix for formula error"""
        suggestions = {
            '#DIV/0!': 'Wrap formula in IFERROR() or check divisor is not zero',
            '#VALUE!': 'Check that all cell references contain expected data types',
            '#REF!': 'Cell reference is broken - check if referenced cells were deleted',
            '#N/A': 'Use IFERROR() or IFNA() to handle missing values',
            '#NAME?': 'Check formula function spelling - might be a typo',
            '#NULL!': 'Check range intersection syntax',
            '#NUM!': 'Check numeric argument is within valid range'
        }
        return suggestions.get(error_value, 'Review formula logic')


def validate_output_file(file_path: str, verbose: bool = False) -> QAReport:
    """
    Validate Excel or PDF output file
    
    Args:
        file_path: Path to file
        verbose: Print detailed report
        
    Returns:
        QAReport with validation results
    """
    validator = QAValidator(verbose=verbose)
    report = validator.validate_file(file_path)
    
    if verbose:
        print(f"\n{'='*80}")
        print(f"📋 QA VALIDATION REPORT - {report.file_type.upper()}")
        print(f"{'='*80}")
        print(f"File: {os.path.basename(file_path)}")
        print(f"Status: {'✅ PASS' if report.is_valid else '❌ FAIL'}")
        print(f"Quality Score: {report.quality_score}/100")
        print(f"\nIssues: {report.critical_issues} critical, {report.warnings} warnings")
        
        if report.summary:
            print(f"\nStatistics:")
            for key, value in report.summary.items():
                print(f"  {key}: {value}")
    
    return report
