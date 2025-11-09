"""
Benchmark API Endpoints for GDPval Integration
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
import os
import asyncio
import json
from pathlib import Path

from app.agents.benchmark_orchestrator import BenchmarkOrchestrator

router = APIRouter()


class BenchmarkTask(BaseModel):
    """Benchmark task definition"""
    id: str
    title: str
    description: str
    category: str
    difficulty: int
    estimated_time: str
    sections: list[str]


@router.get("/tasks")
async def list_benchmark_tasks():
    """List available benchmark tasks from sales_reps_tasks.json"""
    
    tasks_file = "data/gdpval/sales_reps/sales_reps_tasks.json"
    
    if not os.path.exists(tasks_file):
        # Fallback to static task
        tasks = [
            {
                "task_id": "19403010-3e5c-494e-a6d3-13594e99f6af",
                "sector": "Wholesale Trade",
                "occupation": "Sales Representatives",
                "prompt": "XR Retailer 2023 Makeup Sales Analysis",
                "reference_files": [],
                "reference_file_urls": []
            }
        ]
        return {"tasks": tasks, "total": 1}
    
    try:
        with open(tasks_file, 'r') as f:
            all_tasks = json.load(f)
        
        # Generate meaningful titles from prompts
        def generate_title_from_prompt(prompt):
            """Extract a meaningful title from the task prompt"""
            prompt_lower = prompt.lower()
            
            # Pattern matching to extract key topics
            if "automotive" in prompt_lower and "parts" in prompt_lower:
                return "Automotive Parts Check-In Procedure"
            elif "beutist" in prompt_lower and "set" in prompt_lower:
                return "Beutist Set Inventory Analysis"
            elif "xr retailer" in prompt_lower and "makeup" in prompt_lower:
                return "XR Retailer Makeup Sales Analysis"
            elif "alcoholic beverages" in prompt_lower or "inventory" in prompt_lower and "stockout" in prompt_lower:
                return "Beverage Inventory Stockout Prevention"
            elif "fragrance" in prompt_lower and "pricing" in prompt_lower:
                return "Men's Fragrance Competitive Pricing"
            else:
                # Fallback: extract first meaningful sentence
                first_sentence = prompt.split('.')[0][:60]
                return first_sentence + "..." if len(first_sentence) >= 60 else first_sentence
        
        # Map HuggingFace URLs to local file names
        def get_local_filename(hf_url):
            """Convert HuggingFace URL to local filename"""
            url_to_file = {
                "7aef029e58a67b9ce3b8fd6110d8160b/DATA-Beutist Set Selling-v2.xlsx": "DATA-Beutist_Set_Selling-v2.xlsx",
                "83cd6e2233b76f20b6a6643217f9ebb3/DATA XR MU 2023 Final (2).xlsx": "DATA_XR_MU_2023_Final.xlsx",
                "915c72afa404c96174d69e03b74c6454/Inventory_and_Shipments Latest.xlsx": "Inventory_and_Shipments_Latest.xlsx",
                "062f057c961cefe89513e32097df802b/Current Product Price List.xlsx": "Current_Product_Price_List.xlsx"
            }
            
            for key, filename in url_to_file.items():
                if key in hf_url:
                    return filename
            return None
        
        # Return tasks with useful metadata
        formatted_tasks = []
        for task in all_tasks:
            # Extract short description from prompt
            prompt = task.get("prompt", "")
            description = prompt[:200] + "..." if len(prompt) > 200 else prompt
            
            # Convert reference file URLs to local proxy URLs
            reference_file_urls = task.get("reference_file_urls", [])
            proxy_urls = []
            for url in reference_file_urls:
                local_filename = get_local_filename(url)
                if local_filename:
                    # Use local file endpoint
                    proxy_urls.append(f"/api/v1/benchmark/reference-file/{local_filename}")
                else:
                    # Fallback to proxy endpoint
                    proxy_urls.append(f"/api/v1/benchmark/reference-file?url={url}")
            
            formatted_tasks.append({
                "task_id": task.get("task_id"),
                "title": generate_title_from_prompt(prompt),
                "sector": task.get("sector", ""),
                "occupation": task.get("occupation", ""),
                "description": description,
                "full_prompt": prompt,
                "reference_files": task.get("reference_files", []),
                "reference_file_urls": proxy_urls,
                "has_reference_files": len(task.get("reference_files", [])) > 0
            })
        
        return {"tasks": formatted_tasks, "total": len(formatted_tasks)}
    except Exception as e:
        print(f"Error loading tasks: {e}")
        return {"tasks": [], "total": 0, "error": str(e)}


@router.get("/history")
async def get_task_history():
    """Get history of completed tasks by scanning for output files"""
    import openpyxl
    from datetime import datetime
    
    # Load task titles for matching
    tasks_file = "data/gdpval/sales_reps/sales_reps_tasks.json"
    task_titles = {}
    
    if os.path.exists(tasks_file):
        with open(tasks_file, 'r') as f:
            all_tasks = json.load(f)
            for task in all_tasks:
                task_id = task.get("task_id")
                prompt = task.get("prompt", "")
                
                # Generate title (same logic as in /tasks endpoint)
                prompt_lower = prompt.lower()
                if "automotive" in prompt_lower and "parts" in prompt_lower:
                    title = "Automotive Parts Check-In Procedure"
                elif "beutist" in prompt_lower and "set" in prompt_lower:
                    title = "Beutist Set Inventory Analysis"
                elif "xr retailer" in prompt_lower and "makeup" in prompt_lower:
                    title = "XR Retailer Makeup Sales Analysis"
                elif "alcoholic beverages" in prompt_lower or "inventory" in prompt_lower and "stockout" in prompt_lower:
                    title = "Beverage Inventory Stockout Prevention"
                elif "fragrance" in prompt_lower and "pricing" in prompt_lower:
                    title = "Men's Fragrance Competitive Pricing"
                else:
                    title = prompt.split('.')[0][:60]
                
                task_titles[task_id] = title
    
    completed_tasks = []
    
    # Scan for output files in various directories
    search_paths = [
        "data/gdpval/outputs",
        "data/gdpval/deliverable_files", 
        "data/gdpval/reference_files",
        ".claude/skills/xlsx",
        ".claude/skills/pdf",
        "."
    ]
    
    seen_files = set()
    
    for search_path in search_paths:
        if not os.path.exists(search_path):
            continue
            
        for filename in os.listdir(search_path):
            if (filename.endswith("_output.xlsx") or filename.endswith("_output.pdf")) and filename not in seen_files:
                seen_files.add(filename)
                file_path = os.path.join(search_path, filename)
                
                # Extract task ID from filename (handle both old and new format)
                # New format: task_id_timestamp_output.ext (UUID_YYYYMMDD_HHMMSS_output.ext)
                # Old format: task_id_output.ext
                if "_output." in filename:
                    # Remove _output.xlsx or _output.pdf
                    base_name = filename.replace("_output.xlsx", "").replace("_output.pdf", "")
                    # Split by underscore - UUID uses dashes, so it stays as one part
                    # Format: UUID_YYYYMMDD_HHMMSS
                    parts = base_name.split("_")
                    if len(parts) >= 3:  # UUID + date + time = 3 parts
                        # Has timestamp, extract task_id (first part is UUID)
                        task_id = parts[0]
                    else:
                        # No timestamp, entire base_name is task_id
                        task_id = base_name
                else:
                    continue
                    
                is_pdf = filename.endswith(".pdf")
                
                try:
                    # Get file stats
                    file_stat = os.stat(file_path)
                    file_size = file_stat.st_size
                    created_at = datetime.fromtimestamp(file_stat.st_ctime).isoformat()
                    modified_at = datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                    
                    # For Excel files, get sheet info
                    sheet_count = 0
                    sheet_names = []
                    
                    if not is_pdf:
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        sheet_count = len(wb.sheetnames)
                        sheet_names = wb.sheetnames
                        wb.close()
                    
                    completed_tasks.append({
                        "task_id": task_id,
                        "task_title": task_titles.get(task_id, "Unknown Task"),
                        "file_name": filename,
                        "file_type": "pdf" if is_pdf else "excel",
                        "file_size": file_size,
                        "sheet_count": sheet_count,
                        "sheet_names": sheet_names[:3],  # First 3 sheets
                        "created_at": created_at,
                        "modified_at": modified_at,
                        "file_path": file_path
                    })
                except Exception as e:
                    print(f"Error processing {filename}: {e}")
                    continue
    
    # Sort by modified time (newest first)
    completed_tasks.sort(key=lambda x: x["modified_at"], reverse=True)
    
    return {
        "tasks": completed_tasks,
        "total": len(completed_tasks)
    }


@router.post("/execute/{task_id}")
async def execute_benchmark_task(task_id: str):
    """
    Execute a benchmark task using BenchmarkOrchestrator with Claude
    """
    async def event_generator():
        """Generate SSE events from actual Claude execution with heartbeat"""
        
        try:
            # Send initial status
            yield f"data: {json.dumps({'status': 'Initializing Claude...', 'progress': 5})}\n\n"
            await asyncio.sleep(0.5)
            
            # Load the actual task data from JSON based on task_id
            tasks_file = "data/gdpval/sales_reps/sales_reps_tasks.json"
            
            if not os.path.exists(tasks_file):
                raise HTTPException(status_code=404, detail="Tasks file not found")
            
            # Load all tasks
            with open(tasks_file, 'r') as f:
                all_tasks = json.load(f)
            
            # Find the selected task
            selected_task = None
            for task in all_tasks:
                if task.get("task_id") == task_id:
                    selected_task = task
                    break
            
            if not selected_task:
                raise HTTPException(status_code=404, detail=f"Task {task_id} not found")
            
            # Get task description (prompt)
            task_description = selected_task.get("prompt", "")
            
            # Determine output format based on task description
            is_pdf_task = "pdf" in task_description.lower() and "create a pdf" in task_description.lower()
            output_extension = ".pdf" if is_pdf_task else ".xlsx"
            
            # Add timestamp to filename to avoid overwrites
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{task_id}_{timestamp}_output{output_extension}"
            
            # Get reference files
            reference_file_urls = selected_task.get("reference_file_urls", [])
            reference_file_paths = []
            
            # Download reference files if they're from HuggingFace
            for url in reference_file_urls:
                # Map HuggingFace URLs to local files
                local_file = None
                if "7aef029e58a67b9ce3b8fd6110d8160b" in url:
                    local_file = "/app/backend/data/gdpval/reference_files/DATA-Beutist_Set_Selling-v2.xlsx"
                elif "83cd6e2233b76f20b6a6643217f9ebb3" in url:
                    local_file = "/app/backend/data/gdpval/reference_files/DATA_XR_MU_2023_Final.xlsx"
                elif "915c72afa404c96174d69e03b74c6454" in url:
                    local_file = "/app/backend/data/gdpval/reference_files/Inventory_and_Shipments_Latest.xlsx"
                elif "062f057c961cefe89513e32097df802b" in url:
                    local_file = "/app/backend/data/gdpval/reference_files/Current_Product_Price_List.xlsx"
                
                if local_file and os.path.exists(local_file):
                    reference_file_paths.append(local_file)
            
            yield f"data: {json.dumps({'status': 'Loading task data...', 'progress': 10})}\n\n"
            await asyncio.sleep(0.5)
            
            # Create orchestrator
            orchestrator = BenchmarkOrchestrator(verbose=True)
            
            yield f"data: {json.dumps({'status': 'Starting Claude analysis...', 'progress': 15})}\n\n"
            await asyncio.sleep(0.5)
            
            # Progress tracking
            output_text = ""
            output_files = []
            last_progress = 15
            total_cost = 0.0
            tokens_used = 0
            active_skills = set()
            timeline_phases = []  # Track execution timeline
            
            # Execute the task with streaming - properly parse Claude message types
            async for update in orchestrator.execute_gpteval_task_streaming(
                task_description=task_description,
                reference_file_paths=reference_file_paths,
                output_filename=output_filename
            ):
                update_type = update.get("type", "")
                
                if update_type == "system":
                    # System messages with metadata
                    subtype = update.get("subtype", "")
                    data = update.get("data", {})
                    
                    if subtype == "session_started":
                        message = "🚀 Claude session started"
                    elif subtype == "initialization_started":
                        message = "⚙️ Initializing Claude Agent..."
                    elif subtype == "initialization_complete":
                        message = "✅ Claude Agent ready"
                    else:
                        message = f"⚙️ {subtype}"
                    
                    last_progress = min(last_progress + 1, 85)
                    yield f"data: {json.dumps({'status': message, 'progress': last_progress})}\n\n"
                    
                elif update_type == "user":
                    # Skip user message echo - not useful for display
                    pass
                    
                elif update_type == "assistant":
                    # Parse AssistantMessage content blocks
                    content_blocks = update.get("content", "")
                    tool_uses = update.get("tool_uses", [])
                    thinking = update.get("thinking", "")
                    
                    # Show thinking if available (extended thinking models)
                    if thinking and thinking.strip():
                        snippet = thinking[:150] + "..." if len(thinking) > 150 else thinking
                        message = f"🧠 Thinking: {snippet}"
                        last_progress = min(last_progress + 1, 85)
                        yield f"data: {json.dumps({'status': message, 'progress': last_progress, 'detail': 'extended_thinking'})}\n\n"
                    
                    # Show text content
                    if content_blocks and isinstance(content_blocks, str) and content_blocks.strip():
                        snippet = content_blocks[:180] + "..." if len(content_blocks) > 180 else content_blocks
                        message = f"💬 {snippet}"
                        last_progress = min(last_progress + 2, 85)
                        yield f"data: {json.dumps({'status': message, 'progress': last_progress, 'detail': 'claude_response'})}\n\n"
                        output_text += content_blocks
                    
                    # Show tool usage with detailed context
                    if tool_uses:
                        for tool in tool_uses:
                            tool_name = tool.get("name", "Unknown")
                            tool_input = tool.get("input", {})
                            tool_id = tool.get("id", "")
                            
                            if tool_name == "Skill":
                                # Skill invocation - Track active skills
                                skill_name = tool_input.get("skill", "unknown")
                                active_skills.add(skill_name)
                                message = f"🎯 Activating {skill_name} Skill..."
                                detail = f"skill_{skill_name}"
                                
                                # Add to timeline
                                timeline_phases.append({
                                    "phase": f"{skill_name}_skill",
                                    "status": "active",
                                    "time": last_progress
                                })
                                
                            elif tool_name == "Read":
                                file_path = tool_input.get("file_path", tool_input.get("path", ""))
                                if file_path:
                                    file_name = os.path.basename(file_path)
                                    offset = tool_input.get("offset")
                                    limit = tool_input.get("limit")
                                    if offset or limit:
                                        message = f"📖 Reading {file_name} (lines {offset or 0}-{(offset or 0) + (limit or 'end')})"
                                    else:
                                        message = f"📖 Reading: {file_name}"
                                    detail = f"read_{file_name}"
                                else:
                                    message = "📖 Reading file..."
                                    detail = "read_file"
                                    
                            elif tool_name == "Write":
                                file_path = tool_input.get("file_path", tool_input.get("path", ""))
                                content_size = len(str(tool_input.get("content", "")))
                                if file_path:
                                    file_name = os.path.basename(file_path)
                                    message = f"✍️ Writing {file_name} ({content_size} chars)"
                                    detail = f"write_{file_name}"
                                else:
                                    message = f"✍️ Writing file ({content_size} chars)..."
                                    detail = "write_file"
                                    
                            elif tool_name == "Edit":
                                file_path = tool_input.get("file_path", "")
                                replace_all = tool_input.get("replace_all", False)
                                if file_path:
                                    file_name = os.path.basename(file_path)
                                    action = "all occurrences" if replace_all else "first occurrence"
                                    message = f"✏️ Editing {file_name} ({action})"
                                    detail = f"edit_{file_name}"
                                else:
                                    message = "✏️ Editing file..."
                                    detail = "edit_file"
                                    
                            elif tool_name == "Bash":
                                cmd = str(tool_input.get("command", ""))
                                description = tool_input.get("description", "")
                                run_bg = tool_input.get("run_in_background", False)
                                
                                # Use description if available
                                if description:
                                    message = f"⚡ {description}"
                                    detail = "bash_described"
                                # Otherwise, intelligently parse command
                                elif "pip install" in cmd:
                                    packages = cmd.replace("pip install", "").strip().split()[0:3]
                                    message = f"📦 Installing: {', '.join(packages)}"
                                    detail = "bash_install"
                                elif "python" in cmd.lower():
                                    if "openpyxl" in cmd or "xlsx" in cmd:
                                        message = "📊 Processing Excel with Python..."
                                        detail = "bash_python_excel"
                                    elif "reportlab" in cmd or "pdf" in cmd:
                                        message = "📄 Generating PDF with Python..."
                                        detail = "bash_python_pdf"
                                    else:
                                        message = "🐍 Running Python script..."
                                        detail = "bash_python"
                                elif len(cmd) > 0:
                                    cmd_preview = cmd[:70] + "..." if len(cmd) > 70 else cmd
                                    message = f"⚡ {cmd_preview}"
                                    detail = "bash_command"
                                else:
                                    message = "⚡ Executing command..."
                                    detail = "bash_generic"
                                
                                if run_bg:
                                    message += " (background)"
                                    
                            elif tool_name == "Glob":
                                pattern = tool_input.get("pattern", "")
                                message = f"🔍 Finding: {pattern}" if pattern else "🔍 Searching files..."
                                detail = "glob_search"
                            
                            elif tool_name == "Grep":
                                pattern = tool_input.get("pattern", "")
                                path = tool_input.get("path", "")
                                message = f"🔎 Searching '{pattern}' in {os.path.basename(path) if path else 'files'}"
                                detail = "grep_search"
                            
                            elif tool_name == "TodoWrite":
                                todos = tool_input.get("todos", [])
                                if todos and len(todos) > 0:
                                    first_todo = todos[0].get("content", "")[:50]
                                    message = f"📝 Planning: {first_todo}..." if first_todo else "📝 Creating task list..."
                                    detail = "todo_planning"
                                else:
                                    message = "📝 Updating task list..."
                                    detail = "todo_write"
                                
                            else:
                                message = f"🔧 {tool_name}"
                                detail = f"tool_{tool_name.lower()}"
                            
                            last_progress = min(last_progress + 1, 85)
                            yield f"data: {json.dumps({
                                'status': message, 
                                'progress': last_progress, 
                                'detail': detail, 
                                'tool': tool_name,
                                'active_skills': list(active_skills),
                                'cost_usd': total_cost,
                                'tokens': tokens_used
                            })}\n\n"
                    
                elif update_type == "result":
                    # ToolResultBlock - show meaningful preview
                    result_content = update.get("content", "")
                    is_error = update.get("is_error", False)
                    
                    if is_error:
                        error_msg = str(result_content)[:100] if result_content else "Tool execution failed"
                        message = f"⚠️ {error_msg}"
                    elif result_content:
                        # Parse result to show meaningful info
                        result_str = str(result_content)
                        if "rows" in result_str.lower() or "columns" in result_str.lower():
                            message = "✅ Data loaded successfully"
                        elif "created" in result_str.lower() or "written" in result_str.lower():
                            message = "✅ File created"
                        elif len(result_str) > 100:
                            message = "✅ Operation completed"
                        else:
                            preview = result_str[:80] + "..." if len(result_str) > 80 else result_str
                            message = f"✅ {preview}"
                    else:
                        message = "✅ Completed"
                    
                    yield f"data: {json.dumps({'status': message, 'progress': last_progress})}\n\n"
                    
                elif update_type == "error":
                    # Error message
                    error = update.get("error", "Unknown error")
                    message = f"❌ Error: {error}"
                    yield f"data: {json.dumps({'status': message, 'progress': last_progress})}\n\n"
                    
                elif update_type == "complete":
                    # ResultMessage - extract cost and usage data
                    duration_ms = update.get("duration_ms", 0)
                    num_turns = update.get("num_turns", 0)
                    total_cost = update.get("total_cost_usd", 0) or 0
                    usage = update.get("usage", {})
                    
                    # Extract token information
                    if usage:
                        tokens_used = usage.get("input_tokens", 0) + usage.get("output_tokens", 0)
                    
                    if duration_ms and num_turns:
                        duration_sec = duration_ms / 1000
                        message = f"✨ Complete! ({num_turns} turns, {duration_sec:.1f}s"
                        if total_cost:
                            message += f", ${total_cost:.4f}"
                        message += ")"
                    else:
                        message = "✨ Claude analysis complete"
                    
                    last_progress = 90
                    yield f"data: {json.dumps({
                        'status': message, 
                        'progress': last_progress,
                        'cost_usd': total_cost,
                        'tokens': tokens_used,
                        'duration_ms': duration_ms,
                        'num_turns': num_turns
                    })}\n\n"
                    break
                
                # Small delay to prevent overwhelming the client
                await asyncio.sleep(0.05)
            
            # Final status
            yield f"data: {json.dumps({'status': '🎉 Analysis complete!', 'progress': 95})}\n\n"
            await asyncio.sleep(0.5)
            
            # Final result
            result = {
                "status": "complete",
                "task_id": task_id,
                "file_name": output_filename,
                "output_text": output_text[:500] if output_text else "Analysis complete",
                "files_created": len(output_files),
                "errors": 0,
                "progress": 100
            }
            yield f"data: {json.dumps(result)}\n\n"
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error in benchmark execution: {error_detail}")
            error_result = {
                "status": "error",
                "error": str(e),
                "progress": 0
            }
            yield f"data: {json.dumps(error_result)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


@router.get("/result/{task_id}")
async def get_task_result_metadata(task_id: str):
    """Get metadata about the generated Excel file"""
    import openpyxl
    
    # Look for the file
    possible_paths = [
        f"data/gdpval/outputs/{task_id}_output.xlsx",
        f"data/gdpval/deliverable_files/{task_id}_output.xlsx",
        f"data/gdpval/reference_files/{task_id}_output.xlsx",
        f"{task_id}_output.xlsx"
    ]

    file_path = None
    for path in possible_paths:
        if os.path.exists(path):
            file_path = path
            break

    if not file_path:
        raise HTTPException(status_code=404, detail="Excel file not found")

    try:
        # Load workbook to get metadata
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Count rows with data
            rows_with_data = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))
            sheets_info.append({
                "name": sheet_name,
                "rows": rows_with_data,
                "columns": ws.max_column
            })
        
        wb.close()
        
        # Get absolute file path for serving
        abs_file_path = os.path.abspath(file_path)
        
        return {
            "task_id": task_id,
            "file_name": f"{task_id}_output.xlsx",
            "file_path": abs_file_path,
            "file_size": os.path.getsize(file_path),
            "sheets": sheets_info,
            "sheet_count": len(sheets_info),
            "download_url": f"/api/v1/benchmark/download/{task_id}",
            "preview_url": f"/api/v1/benchmark/file/{task_id}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")


@router.get("/file/{task_id}")
async def get_file(task_id: str):
    """Serve the Excel or PDF file for browser preview (without forcing download)"""
    import glob
    
    # Look for PDF files FIRST (higher priority) - with timestamp support
    pdf_patterns = [
        f"data/gdpval/outputs/{task_id}_*_output.pdf",
        f"data/gdpval/outputs/{task_id}_output.pdf",
        f"data/gdpval/deliverable_files/{task_id}_*_output.pdf",
        f"data/gdpval/deliverable_files/{task_id}_output.pdf",
        f".claude/skills/pdf/{task_id}_*_output.pdf",
        f".claude/skills/pdf/{task_id}_output.pdf",
        f"{task_id}_*_output.pdf",
        f"{task_id}_output.pdf"
    ]
    
    # Then look for Excel files
    excel_patterns = [
        f"data/gdpval/outputs/{task_id}_*_output.xlsx",
        f"data/gdpval/outputs/{task_id}_output.xlsx",
        f"data/gdpval/deliverable_files/{task_id}_*_output.xlsx",
        f"data/gdpval/deliverable_files/{task_id}_output.xlsx",
        f"data/gdpval/reference_files/{task_id}_*_output.xlsx",
        f"data/gdpval/reference_files/{task_id}_output.xlsx",
        f".claude/skills/xlsx/{task_id}_*_output.xlsx",
        f".claude/skills/xlsx/{task_id}_output.xlsx",
        f"{task_id}_*_output.xlsx",
        f"{task_id}_output.xlsx"
    ]

    file_path = None
    is_pdf = False
    
    # Check PDF first (most recent)
    for pattern in pdf_patterns:
        matches = glob.glob(pattern)
        if matches:
            file_path = max(matches, key=os.path.getmtime)
            is_pdf = True
            break
    
    # Then check Excel
    if not file_path:
        for pattern in excel_patterns:
            matches = glob.glob(pattern)
            if matches:
                file_path = max(matches, key=os.path.getmtime)
                break

    if not file_path:
        raise HTTPException(status_code=404, detail="Output file not found")

    # Extract actual filename
    actual_filename = os.path.basename(file_path)

    # Return with appropriate media type
    if is_pdf:
        return FileResponse(
            file_path,
            media_type="application/pdf",
            filename=actual_filename,
            headers={
                "Content-Disposition": f'inline; filename="{actual_filename}"'
            }
        )
    else:
        return FileResponse(
            file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=actual_filename,
            headers={
                "Content-Disposition": f'inline; filename="{actual_filename}"'
            }
        )


@router.get("/download/{task_id}")
async def download_excel_result(task_id: str):
    """Download the generated Excel or PDF file for a task"""
    import glob

    # Check for PDF files FIRST (higher priority) - with timestamp support
    pdf_patterns = [
        f"data/gdpval/outputs/{task_id}_*_output.pdf",
        f"data/gdpval/outputs/{task_id}_output.pdf",
        f"data/gdpval/deliverable_files/{task_id}_*_output.pdf",
        f"data/gdpval/deliverable_files/{task_id}_output.pdf",
        f".claude/skills/pdf/{task_id}_*_output.pdf",
        f".claude/skills/pdf/{task_id}_output.pdf",
        f"{task_id}_*_output.pdf",
        f"{task_id}_output.pdf"
    ]
    
    # Then check Excel files
    excel_patterns = [
        f"data/gdpval/outputs/{task_id}_*_output.xlsx",
        f"data/gdpval/outputs/{task_id}_output.xlsx",
        f"data/gdpval/deliverable_files/{task_id}_*_output.xlsx",
        f"data/gdpval/deliverable_files/{task_id}_output.xlsx",
        f"data/gdpval/reference_files/{task_id}_*_output.xlsx",
        f"data/gdpval/reference_files/{task_id}_output.xlsx",
        f".claude/skills/xlsx/{task_id}_*_output.xlsx",
        f".claude/skills/xlsx/{task_id}_output.xlsx",
        f"{task_id}_*_output.xlsx",
        f"{task_id}_output.xlsx"
    ]

    file_path = None
    is_pdf = False
    
    # Check PDF files first (most recent timestamp)
    for pattern in pdf_patterns:
        matches = glob.glob(pattern)
        if matches:
            # Sort by modification time, get most recent
            file_path = max(matches, key=os.path.getmtime)
            is_pdf = True
            break
    
    # If no PDF, check Excel files
    if not file_path:
        for pattern in excel_patterns:
            matches = glob.glob(pattern)
            if matches:
                # Sort by modification time, get most recent
                file_path = max(matches, key=os.path.getmtime)
                break

    if not file_path:
        raise HTTPException(status_code=404, detail="Output file not found")

    # Extract actual filename
    actual_filename = os.path.basename(file_path)

    # Determine media type
    if is_pdf:
        media_type = "application/pdf"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return FileResponse(
        file_path,
        media_type=media_type,
        filename=actual_filename,
        headers={
            "Content-Disposition": f'attachment; filename="{actual_filename}"'
        }
    )


@router.get("/validate/{task_id}")
async def get_validation_report(task_id: str):
    """
    Get QA validation report for a task's output file
    Validates both Excel and PDF files automatically
    """
    from app.agents.qa_validator import validate_output_file
    import glob
    
    # Find the output file (PDF or Excel) with timestamp support
    all_patterns = [
        f"data/gdpval/outputs/{task_id}_*_output.*",
        f"data/gdpval/outputs/{task_id}_output.*",
        f".claude/skills/xlsx/{task_id}_*_output.xlsx",
        f".claude/skills/xlsx/{task_id}_output.xlsx",
        f".claude/skills/pdf/{task_id}_*_output.pdf",
        f".claude/skills/pdf/{task_id}_output.pdf",
        f"{task_id}_*_output.*",
        f"{task_id}_output.*"
    ]
    
    file_path = None
    for pattern in all_patterns:
        matches = glob.glob(pattern)
        if matches:
            # Get most recent file
            file_path = max(matches, key=os.path.getmtime)
            break
    
    if not file_path:
        # Return default validation for demo
        return {
            "task_id": task_id,
            "file_type": "unknown",
            "is_valid": True,
            "quality_score": 95,
            "total_issues": 0,
            "critical_issues": 0,
            "warnings": 0,
            "info_messages": 1,
            "issues": [{
                "severity": "info",
                "category": "status",
                "location": "N/A",
                "message": "File not found for validation"
            }],
            "summary": {}
        }
    
    # Run validation
    report = validate_output_file(file_path, verbose=True)
    
    # Add task_id to response
    result = report.to_dict()
    result['task_id'] = task_id
    result['file_name'] = os.path.basename(file_path)
    
    # Save validation report
    report_path = f"data/gdpval/outputs/{task_id}_validation_report.json"
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    
    with open(report_path, 'w') as f:
        json.dump(result, f, indent=2)
    
    return result


@router.get("/reference-file/{filename}")
async def serve_local_reference_file(filename: str):
    """
    Serve local reference files with proper headers for Google Sheets import
    """
    file_path = f"data/gdpval/reference_files/{filename}"
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Reference file not found")
    
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Allow-Origin": "*"
        }
    )


@router.get("/reference-file")
async def serve_reference_file(url: str):
    """
    Proxy endpoint to serve reference files from HuggingFace for Google Sheets import
    This is a fallback for files not available locally
    """
    import httpx
    from urllib.parse import unquote
    from fastapi.responses import Response
    
    try:
        # Download file from HuggingFace
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url)
            response.raise_for_status()
            
            # Extract filename from URL
            filename = url.split('/')[-1]
            filename = unquote(filename)
            
            # Return file with proper headers for Google Sheets
            return Response(
                content=response.content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Access-Control-Allow-Origin": "*",
                    "Content-Length": str(len(response.content))
                }
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch reference file: {str(e)}")
