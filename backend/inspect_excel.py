#!/usr/bin/env python3
"""
Inspect Excel file for DIV/0 errors and other issues
"""
import sys
from openpyxl import load_workbook


def inspect_excel(file_path):
    """Inspect an Excel file and print all cells with their values and formulas"""

    print(f"Inspecting: {file_path}\n")

    # Load both versions
    print("Loading workbook with formulas...")
    wb_formula = load_workbook(file_path, data_only=False)

    print("Loading workbook with calculated values...")
    wb_data = load_workbook(file_path, data_only=True)

    for sheet_name in wb_formula.sheetnames:
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*80}\n")

        sheet_formula = wb_formula[sheet_name]
        sheet_data = wb_data[sheet_name]

        # Find cells with errors
        errors_found = False

        for row_idx, row in enumerate(sheet_formula.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                data_cell = sheet_data.cell(row=row_idx, column=col_idx)

                # Check for errors in calculated value
                if data_cell.value and isinstance(data_cell.value, str):
                    if '#' in str(data_cell.value):
                        errors_found = True
                        formula = cell.value if str(cell.value).startswith('=') else 'N/A'
                        print(f"ERROR at {cell.coordinate}:")
                        print(f"  Value: {data_cell.value}")
                        print(f"  Formula: {formula}")
                        print()

                # Also check cell value directly
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('#'):
                        if not (data_cell.value and isinstance(data_cell.value, str) and '#' in data_cell.value):
                            errors_found = True
                            print(f"ERROR at {cell.coordinate}:")
                            print(f"  Value: {cell.value}")
                            print(f"  Formula: {cell.value if str(cell.value).startswith('=') else 'N/A'}")
                            print()

        if not errors_found:
            print("✅ No errors found in this sheet")

    wb_formula.close()
    wb_data.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python inspect_excel.py <excel_file>")
        sys.exit(1)

    inspect_excel(sys.argv[1])
